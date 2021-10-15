import  openpyxl

# Criar uma planila (book)
book = openpyxl.Workbook()
# como visualizar paginas existentes.
print(book.sheetnames)

# como criar uma pagina
book.create_sheet('Frutas')

# como selecionar uma pagina
paginas_frutas = book['Frutas']

# adiconar linhas, passa uma lista
paginas_frutas.append(['Banana','5','R$3,90'])
paginas_frutas.append(['Maça','15','R$5,90'])
paginas_frutas.append(['Abacate','25','R$7,00'])
paginas_frutas.append(['Tangerina','25','R$10.00'])
paginas_frutas.append(['Morango','15','R$8,90'])

# salvar planilha
book.save('Planilha de Compras.xlsx')
